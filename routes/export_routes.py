"""
Excel export API route — generates .xlsx files with person & criminal record data.
Adapted for Firebase Firestore.
"""
import io
import os
from datetime import datetime
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from google.cloud import firestore
from database.database import get_db
from auth.auth import get_current_user, require_role

router = APIRouter(prefix="/api/export", tags=["Export"])

@router.get("/excel")
def export_to_excel(
    current_user: dict = Depends(require_role("admin", "officer")),
    db: firestore.Client = Depends(get_db),
):
    """
    Export all persons and their criminal records to an Excel file.
    Returns a downloadable .xlsx file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        from fastapi.responses import JSONResponse
        return JSONResponse(
            status_code=500,
            content={"detail": "openpyxl not installed. Run: pip install openpyxl"}
        )

    wb = Workbook()

    # ── Sheet 1: Persons ──────────────────────────────────────────
    ws_persons = wb.active
    ws_persons.title = "Persons"

    # Styling
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    risk_fills = {
        "High": PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid"),
        "Medium": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "Low": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
    }

    person_headers = [
        "Person ID", "Full Name", "Date of Birth", "Gender", "Nationality",
        "Address", "Government ID", "Record Status", "Risk Level",
        "Has Face Data", "Created At", "Updated At"
    ]
    for col, header in enumerate(person_headers, 1):
        cell = ws_persons.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Fetch Data
    persons_docs = list(db.collection("persons").stream())
    persons = [p.to_dict() for p in persons_docs]
    persons.sort(key=lambda x: x.get("id", 0))

    # Caching for Records sheet later
    person_names = {p.get("id"): p.get("full_name", "Unknown") for p in persons}
    
    for row_idx, p in enumerate(persons, 2):
        data = [
            p.get("id"), p.get("full_name") or "N/A", p.get("date_of_birth") or "N/A", p.get("gender") or "N/A",
            p.get("nationality") or "N/A", p.get("address") or "N/A", p.get("government_id_number") or "N/A",
            p.get("record_status") or "N/A", p.get("risk_level") or "N/A",
            "Yes" if p.get("face_embedding_encrypted") else "No",
            p.get("created_at") or "N/A",
            p.get("updated_at") or "N/A",
        ]
        for col, value in enumerate(data, 1):
            cell = ws_persons.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            if col == 9 and value in risk_fills:
                cell.fill = risk_fills[value]
                cell.font = Font(bold=True)

    for col in range(1, len(person_headers) + 1):
        max_len = max(
            len(str(ws_persons.cell(row=r, column=col).value or ""))
            for r in range(1, ws_persons.max_row + 1)
        )
        ws_persons.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = min(max_len + 4, 40)

    # ── Sheet 2: Criminal Records ─────────────────────────────────
    ws_records = wb.create_sheet("Criminal Records")

    record_headers = [
        "Record ID", "Person ID", "Person Name", "Crime Type", "Crime Description",
        "Case Number", "Date of Offense", "Arrest Date", "Conviction Status",
        "Sentence Details", "Law Enforcement Agency", "Court Name",
        "Officer Notes", "Last Updated"
    ]

    for col, header in enumerate(record_headers, 1):
        cell = ws_records.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    records_docs = list(db.collection("criminal_records").stream())
    records = [r.to_dict() for r in records_docs]
    records.sort(key=lambda x: x.get("id", 0))

    for row_idx, r in enumerate(records, 2):
        pid = r.get("person_id")
        pname = person_names.get(pid, "Unknown")
        data = [
            r.get("id"), pid, pname,
            r.get("crime_type") or "N/A", r.get("crime_description") or "N/A",
            r.get("case_number") or "N/A", r.get("date_of_offense") or "N/A",
            r.get("arrest_date") or "N/A", r.get("conviction_status") or "N/A",
            r.get("sentence_details") or "N/A", r.get("law_enforcement_agency") or "N/A",
            r.get("court_name") or "N/A", r.get("officer_notes") or "N/A",
            r.get("last_updated") or "N/A",
        ]
        for col, value in enumerate(data, 1):
            cell = ws_records.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # ── Sheet 3: Summary ──────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    ws_summary.cell(row=1, column=1, value="Criminal Recognition System - Export Summary").font = Font(bold=True, size=14)
    ws_summary.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws_summary.cell(row=3, column=1, value=f"Generated by: {current_user.get('full_name', 'Unknown')}")
    ws_summary.cell(row=5, column=1, value="Total Persons:").font = Font(bold=True)
    ws_summary.cell(row=5, column=2, value=len(persons))
    ws_summary.cell(row=6, column=1, value="Total Criminal Records:").font = Font(bold=True)
    ws_summary.cell(row=6, column=2, value=len(records))

    ws_summary.cell(row=8, column=1, value="By Record Status:").font = Font(bold=True)
    status_counts = {}
    for p in persons:
        s = p.get("record_status") or "Unknown"
        status_counts[s] = status_counts.get(s, 0) + 1
        
    for i, (status, count) in enumerate(status_counts.items()):
        ws_summary.cell(row=9 + i, column=1, value=status)
        ws_summary.cell(row=9 + i, column=2, value=count)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"criminal_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
